import openpyxl
from datetime import datetime
from threading import Thread


class StationProvider(Thread):

    # daemon = True
    train_capacity = [2000, 800, 1200]
    column_names = ['Время', 'Событие', 'Пришли', 'Зал ожидания', 'Платформа 1', 'Платформа 2', 'Платформа 3',
                    'Поезд 1', 'Поезд 2', 'Поезд 3']

    def __init__(self, gui_mod):
        Thread.__init__(self, target=self.main_loop)
        self.excel_file = openpyxl.Workbook()
        self.sheet = self.excel_file.create_sheet('Sheet_1', 0)
        for i in range(1, 11):
            self.sheet.cell(row=1, column=i).value = self.column_names[i - 1]
        self.row_num = 2
        self.gui_mod = gui_mod
        self.alive = True
        self.people_dict = {
            'waiting_hall': 0,
            'platform_1': 0,
            'platform_2': 0,
            'platform_3': 0,
            'train_1': 0,
            'train_2': 0,
            'train_3': 0,
            'people_from_train_1': 0,
            'people_from_train_2': 0,
            'people_from_train_3': 0
        }
        self.arriving = False           # поезд приехал
        self.boarding = False           # началась посадка
        self.no_train = True            # поезд уехал
        self.people_came = False        # пришли новые люди
        self.people_num = 0             # их количество

        self.first = True
        self.left = False
        self.pause = True

        self.people_time = False
        self.hours = 0
        self.minutes = 0
        self.speed = 1 / 100            # 100 менять на нужную

    def update_info(self, message):
        min_ = str(self.minutes)
        if len(min_) == 1:
            min_ = '0' + min_
        time = ''.join([str(self.hours), ':', min_])
        self.gui_mod.time_label.config(text=time)
        self.gui_mod.info_label.config(text=message)
        if self.gui_mod.file_var.get():
            self.sheet.cell(row=self.row_num, column=1).value = time
            self.sheet.cell(row=self.row_num, column=2).value = message

    def save_station_state(self):
        self.sheet.cell(row=self.row_num, column=4).value = self.people_dict['waiting_hall']
        self.sheet.cell(row=self.row_num, column=5).value = self.people_dict['platform_1']
        self.sheet.cell(row=self.row_num, column=6).value = self.people_dict['platform_2']
        self.sheet.cell(row=self.row_num, column=7).value = self.people_dict['platform_3']
        self.sheet.cell(row=self.row_num, column=8).value = self.people_dict['train_1']
        self.sheet.cell(row=self.row_num, column=9).value = self.people_dict['train_2']
        self.sheet.cell(row=self.row_num, column=10).value = self.people_dict['train_3']

    def main_loop(self):
        while self.alive:
            if self.people_came and self.alive:
                self.people_came_func()
                self.people_came = False
            if self.arriving and self.alive:
                self.arriving_func()
                self.arriving = False
                self.first = True
            if self.boarding and self.alive:
                self.boarding_func()
            if self.left and self.alive:
                self.people_dict['train_1'] = 0
                self.people_dict['train_2'] = 0
                self.people_dict['train_3'] = 0
                print(self.hours, ":", self.minutes,
                      '\nПоезд уехал')
                self.update_info('Поезд уехал')
                self.gui_mod.train_1.config(text='', bg='lavender')
                self.gui_mod.train_2.config(text='', bg='lavender')
                self.gui_mod.train_3.config(text='', bg='lavender')

                if self.gui_mod.file_var.get():
                    self.save_station_state()
                    self.row_num += 1

                self.left = False

    def people_came_func(self):
        # обработка новых людей
        if self.gui_mod.file_var.get():
            self.sheet.cell(row=self.row_num, column=3).value = self.people_num
        if self.no_train and self.alive:
            wh = round(self.people_num * 0.66)
            p_to_platform = self.people_num - wh
            pl_1 = round(p_to_platform * 0.5)
            pl_2 = round(p_to_platform * 0.2)
            self.people_dict['waiting_hall'] += wh
            self.people_dict['platform_1'] += pl_1
            self.people_dict['platform_2'] += pl_2
            self.people_dict['platform_3'] += p_to_platform - pl_1 - pl_2
        elif self.alive:
            pl_1 = round(self.people_num * 0.5)
            pl_2 = round(self.people_num * 0.2)
            self.people_dict['platform_1'] += pl_1
            self.people_dict['platform_2'] += pl_2
            self.people_dict['platform_3'] += self.people_num - pl_1 - pl_2

        # to gui
        self.update_info('Пришли люди')
        self.gui_mod.wait_num.config(text=str(self.people_dict['waiting_hall']))
        self.gui_mod.pl_1.config(text=str(self.people_dict['platform_1']))
        self.gui_mod.pl_2.config(text=str(self.people_dict['platform_2']))
        self.gui_mod.pl_3.config(text=str(self.people_dict['platform_3']))

        if self.gui_mod.file_var.get():
            self.save_station_state()
            self.row_num += 1

        print('\nПришли люди. Загруженность вокзала:'
              '\n     - зал ожидания: ', self.people_dict['waiting_hall'],
              '\n     - платформа 1: ', self.people_dict['platform_1'],
              '\n     - платформа 2: ', self.people_dict['platform_2'],
              '\n     - платформа 3: ', self.people_dict['platform_3']
              )

    def arriving_func(self):
        # to gui
        self.update_info('Приехали поезда')
        self.gui_mod.wait_num.config(text=str(self.people_dict['waiting_hall']))
        self.gui_mod.pl_1.config(text=str(self.people_dict['platform_1']))
        self.gui_mod.pl_2.config(text=str(self.people_dict['platform_2']))
        self.gui_mod.pl_3.config(text=str(self.people_dict['platform_3']))
        self.gui_mod.train_1.config(text=str(self.people_dict['train_1']), bg='navy', fg='snow')
        self.gui_mod.train_2.config(text=str(self.people_dict['train_2']), bg='navy', fg='snow')
        self.gui_mod.train_3.config(text=str(self.people_dict['train_3']), bg='navy', fg='snow')

        if self.gui_mod.file_var.get():
            self.save_station_state()
            self.row_num += 1

        print(self.hours, ":", self.minutes,
              '\nПриехали поезда:'
              '\n     - поезд 1: ', self.people_dict['train_1'],
              '\n     - поезд 2: ', self.people_dict['train_2'],
              '\n     - поезд 3: ', self.people_dict['train_3']
              )
        print('Загруженность вокзала:'
              '\n     - зал ожидания: ', self.people_dict['waiting_hall'],
              '\n     - платформа 1: ', self.people_dict['platform_1'],
              '\n     - платформа 2: ', self.people_dict['platform_2'],
              '\n     - платформа 3: ', self.people_dict['platform_3']
              )

    def boarding_func(self):
        if self.first and self.alive:
            for i in range(0, 3):
                # приехавшие люди уходят
                if not self.alive:
                    break
                self.people_dict[''.join(['train_', str(i + 1)])] = 0

            # to gui
            min_ = str(self.minutes)
            if len(min_) == 1:
                min_ = '0' + min_
            time = ''.join([str(self.hours), ':', min_])
            self.gui_mod.time_label.config(text=time)
            self.gui_mod.wait_num.config(text=str(self.people_dict['waiting_hall']))
            self.gui_mod.pl_1.config(text=str(self.people_dict['platform_1']))
            self.gui_mod.pl_2.config(text=str(self.people_dict['platform_2']))
            self.gui_mod.pl_3.config(text=str(self.people_dict['platform_3']))

            print(self.hours, ":", self.minutes,
                  '\nЛюди из поездов ушли. Загруженность вокзала:'
                  '\n     - зал ожидания: ', self.people_dict['waiting_hall'],
                  '\n     - платформа 1: ', self.people_dict['platform_1'],
                  '\n     - платформа 2: ', self.people_dict['platform_2'],
                  '\n     - платформа 3: ', self.people_dict['platform_3'],
                  '\n     - поезд 1: ', self.people_dict['train_1'],
                  '\n     - поезд 2: ', self.people_dict['train_2'],
                  '\n     - поезд 3: ', self.people_dict['train_3']
                  )

            if self.gui_mod.file_var.get():
                self.sheet.cell(row=self.row_num, column=1).value = time
                self.sheet.cell(row=self.row_num, column=2).value = 'Люди из поездов ушли'
                self.save_station_state()
                self.row_num += 1

            # люди из зала идут на платформы
            if self.people_dict['waiting_hall'] != 0 and self.alive:
                pl_1 = round(self.people_dict['waiting_hall'] * 0.5)
                pl_2 = round(self.people_dict['waiting_hall'] * 0.2)
                self.people_dict['platform_1'] += pl_1
                self.people_dict['platform_2'] += pl_2
                self.people_dict['platform_3'] += self.people_dict['waiting_hall'] - pl_1 - pl_2
                self.people_dict['waiting_hall'] = 0
            self.first = False

            self.gui_mod.wait_num.config(text=str(self.people_dict['waiting_hall']))
            self.gui_mod.pl_1.config(text=str(self.people_dict['platform_1']))
            self.gui_mod.pl_2.config(text=str(self.people_dict['platform_2']))
            self.gui_mod.pl_3.config(text=str(self.people_dict['platform_3']))

            print(self.hours, ":", self.minutes,
                  '\nЛюди из зала ожидания пошли на платформы. Загруженность вокзала:'
                  '\n     - зал ожидания: ', self.people_dict['waiting_hall'],
                  '\n     - платформа 1: ', self.people_dict['platform_1'],
                  '\n     - платформа 2: ', self.people_dict['platform_2'],
                  '\n     - платформа 3: ', self.people_dict['platform_3']
                  )

            if self.gui_mod.file_var.get():
                self.sheet.cell(row=self.row_num, column=1).value = time
                self.sheet.cell(row=self.row_num, column=2).value = 'Люди из зала ожидания пошли на платформы'
                self.save_station_state()
                self.row_num += 1

            # люди садятся в поезд
        for i in range(0, 3):
            if not self.alive:
                break
            if self.people_dict[''.join(['platform_', str(i + 1)])] != 0 and self.alive:
                free = self.train_capacity[i] - self.people_dict[''.join(['train_', str(i + 1)])]
                if free != 0 and self.alive:
                    if self.people_dict[''.join(['platform_', str(i + 1)])] <= free and self.alive:
                        self.people_dict[''.join(['train_', str(i + 1)])] += \
                            self.people_dict[''.join(['platform_', str(i + 1)])]
                        self.people_dict[''.join(['platform_', str(i + 1)])] = 0
                    elif self.alive:
                        self.people_dict[''.join(['train_', str(i + 1)])] += free
                        self.people_dict[''.join(['platform_', str(i + 1)])] -= free
                    self.update_info('Посадка')
                    self.gui_mod.wait_num.config(text=str(self.people_dict['waiting_hall']))
                    self.gui_mod.pl_1.config(text=str(self.people_dict['platform_1']))
                    self.gui_mod.pl_2.config(text=str(self.people_dict['platform_2']))
                    self.gui_mod.pl_3.config(text=str(self.people_dict['platform_3']))
                    self.gui_mod.train_1.config(text=str(self.people_dict['train_1']))
                    self.gui_mod.train_2.config(text=str(self.people_dict['train_2']))
                    self.gui_mod.train_3.config(text=str(self.people_dict['train_3']))

                    if self.gui_mod.file_var.get():
                        self.save_station_state()
                        self.row_num += 1

                    print('Посадка. Загруженность вокзала:'
                          '\n     - платформа 1: ', self.people_dict['platform_1'],
                          '\n     - платформа 2: ', self.people_dict['platform_2'],
                          '\n     - платформа 3: ', self.people_dict['platform_3'],
                          '\n     - поезд 1: ', self.people_dict['train_1'],
                          '\n     - поезд 2: ', self.people_dict['train_2'],
                          '\n     - поезд 3: ', self.people_dict['train_3']
                          )

    def quit(self):
        self.alive = False
        self.excel_file.save(''.join(['data/', datetime.strftime(datetime.today(), "%H-%M-%S"), '.xlsx']))
